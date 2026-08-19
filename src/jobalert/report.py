from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from jobalert.models import Job

HEADERS = [
    "Job Title",
    "Company",
    "Location",
    "Experience",
    "Skills",
    "Posted Time",
    "Email Received",
    "Source",
    "Job URL",
    "Match Score",
    "Priority",
    "Date Verified",
    "Status",
    "Reason",
    "Unique ID",
]


def _row(job: Job) -> list[object]:
    return [
        job.title,
        job.company,
        job.location,
        job.experience,
        ", ".join(job.skills),
        job.posted_at.isoformat() if job.posted_at else "Unknown",
        job.email_received_at.isoformat(),
        job.source,
        job.url,
        job.score,
        job.priority,
        "Yes" if job.date_verified else "No",
        "New",
        job.reason,
        job.unique_id,
    ]


def write_report(jobs: list[Job], output: Path) -> Path:
    output.mkdir(parents=True, exist_ok=True)
    path = output / f"daily-jobs-{datetime.now().date().isoformat()}.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in ("High Priority", "Medium Priority", "Needs Review", "All Jobs"):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(HEADERS)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        selected = (
            jobs
            if sheet_name == "All Jobs"
            else [job for job in jobs if job.priority == sheet_name]
        )
        for job in selected:
            sheet.append(_row(job))
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(
                45, max(12, max(len(str(cell.value or "")) for cell in column) + 2)
            )
    summary = workbook.create_sheet("Run Summary")
    counts = Counter(job.priority for job in jobs)
    for row in (
        ("Metric", "Value"),
        ("New jobs", len(jobs)),
        ("High priority", counts["High Priority"]),
        ("Medium priority", counts["Medium Priority"]),
        ("Needs review", counts["Needs Review"]),
    ):
        summary.append(row)
    workbook.save(path)
    return path
