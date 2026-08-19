from datetime import UTC, datetime
from email.message import EmailMessage
from pathlib import Path

from openpyxl import load_workbook

from jobalert.config import Config
from jobalert.parser import parse_message
from jobalert.report import write_report
from jobalert.scoring import score_job


def config() -> Config:
    return Config(
        24,
        60,
        80,
        5,
        ["remote"],
        ["devops", "sre"],
        ["aws", "docker", "terraform"],
        ["linkedin.com"],
    )


def message() -> EmailMessage:
    mail = EmailMessage()
    mail["From"] = "jobs-noreply@linkedin.com"
    mail["Date"] = "Wed, 19 Aug 2026 08:00:00 +0000"
    mail.set_content("Job alert")
    mail.add_alternative(
        '<a href="https://example.com/job/1" data-company="Acme" data-location="Remote">'
        "DevOps Engineer</a><p>AWS Docker Terraform, 2-4 years, 3 hours ago</p>",
        subtype="html",
    )
    return mail


def test_parse_score_and_report(tmp_path: Path) -> None:
    jobs = parse_message(message())
    assert len(jobs) == 1
    job = score_job(jobs[0], config(), datetime(2026, 8, 19, 9, tzinfo=UTC))
    assert job.source == "LinkedIn"
    assert job.score == 100
    assert job.priority == "High Priority"
    report = write_report([job], tmp_path)
    workbook = load_workbook(report)
    assert workbook["High Priority"].max_row == 2
